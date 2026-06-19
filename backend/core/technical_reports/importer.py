from __future__ import annotations

import csv
import io
import re
import unicodedata
from datetime import datetime
from typing import Any

from openpyxl import load_workbook

from backend.core.technical_reports.models import (
    TechnicalReport,
    _safe_int,
    _safe_str,
    create_empty_report,
    report_id_from_number,
)

MESES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}

COLUMN_MAPPING = {
    "nroinforme": "informe_id",
    "numeroinforme": "informe_id",
    "informeid": "informe_id",
    "informe": "informe_id",
    "id": "informe_id",
    "item": "informe_id",
    "centrodeservicio": "cs",
    "centroservicio": "cs",
    "cs": "cs",
    "sede": "cs",
    "localidad": "cs",
    "contratista": "contratista",
    "sgio": "sgio",
    "codigoinfraestructura": "codigo_infraestructura",
    "codinfraestructura": "codigo_infraestructura",
    "infraestructura": "codigo_infraestructura",
    "codigo": "codigo_infraestructura",
    "ubicacion": "ubicacion",
    "direccion": "ubicacion",
    "suministro": "suministro",
    "nrosuministro": "suministro",
    "nis": "suministro",
    "tipo": "tipo",
    "tipoestructura": "tipo",
    "volumen": "volumen",
    "volumenm3": "volumen",
    "capacidad": "volumen",
    "dia": "dia",
    "mes": "mes",
    "ano": "anio",
    "año": "anio",
    "anio": "anio",
    "cajaregistro": "caja_registro",
    "cajaderegistro": "caja_registro",
    "marcotapa": "marco_tapa",
    "marcoytapa": "marco_tapa",
    "escalerainterior": "escalera_interior",
    "escaleraint": "escalera_interior",
    "escaleraexterior": "escalera_exterior",
    "escaleraext": "escalera_exterior",
    "cubainterior": "cuba_interior",
    "cubaint": "cuba_interior",
    "cubaexterior": "cuba_exterior",
    "cubaext": "cuba_exterior",
    "lozafondo": "loza_fondo",
    "lozadefondo": "loza_fondo",
    "lozatechointerior": "loza_techo_interior",
    "lozatechoint": "loza_techo_interior",
    "lozatechoexterior": "loza_techo_exterior",
    "lozatechoext": "loza_techo_exterior",
    "ductoventilacion": "ducto_ventilacion",
    "ductodeventilacion": "ducto_ventilacion",
    "ventilacion": "ducto_ventilacion",
    "cercoperimetrico": "cerco_perimetrico",
    "cerco": "cerco_perimetrico",
    "descarga": "descarga",
    "tuberiadescarga": "descarga",
    "diametro": "medidas_diametro",
    "diametrointerno": "medidas_diametro_interno",
    "alturautil": "medidas_altura_util",
    "alturatotal": "medidas_altura_total",
    "observaciones": "observaciones",
    "observacion": "observaciones",
    "sugerencias": "sugerencias",
    "sugerencia": "sugerencias",
    "recomendaciones": "sugerencias",
    "recomendacion": "sugerencias",
}

for prefix, target in [
    ("obscajaregistro", "obs_caja_registro"),
    ("sugcajaregistro", "sug_caja_registro"),
    ("obsmarcotapa", "obs_marco_tapa"),
    ("sugmarcotapa", "sug_marco_tapa"),
    ("obsescalerainterior", "obs_escalera_int"),
    ("sugescalerainterior", "sug_escalera_int"),
    ("obsescaleraexterior", "obs_escalera_ext"),
    ("sugescaleraexterior", "sug_escalera_ext"),
    ("obscubainterior", "obs_cuba_int"),
    ("sugcubainterior", "sug_cuba_int"),
    ("obscubaexterior", "obs_cuba_ext"),
    ("sugcubaexterior", "sug_cuba_ext"),
    ("obslozafondo", "obs_loza_fondo"),
    ("suglozafondo", "sug_loza_fondo"),
    ("obslozatechoint", "obs_loza_techo_int"),
    ("suglozatechoint", "sug_loza_techo_int"),
    ("obslozatechoext", "obs_loza_techo_ext"),
    ("suglozatechoext", "sug_loza_techo_ext"),
    ("obsducto", "obs_ducto"),
    ("sugducto", "sug_ducto"),
    ("obscerco", "obs_cerco"),
    ("sugcerco", "sug_cerco"),
    ("obsdescarga", "obs_descarga"),
    ("sugdescarga", "sug_descarga"),
]:
    COLUMN_MAPPING[prefix] = target

for group in ["conduccion", "impulsion", "aduccion", "bypass", "desague"]:
    for diameter in ["2", "3", "4", "6", "8", "10", "12"]:
        COLUMN_MAPPING[f"valvulas{group}{diameter}"] = f"valvulas_{group}_{diameter}"
        COLUMN_MAPPING[f"valv{group}{diameter}"] = f"valvulas_{group}_{diameter}"
    COLUMN_MAPPING[f"obsvalvulas{group}"] = f"obs_valvulas_{group}"
    COLUMN_MAPPING[f"sugvalvulas{group}"] = f"sug_valvulas_{group}"

for group in ["aduccion", "succion", "desague"]:
    for diameter in ["2", "3", "4", "6", "8", "10", "14"]:
        COLUMN_MAPPING[f"canastillas{group}{diameter}"] = f"canastillas_{group}_{diameter}"
        COLUMN_MAPPING[f"canast{group}{diameter}"] = f"canastillas_{group}_{diameter}"
    COLUMN_MAPPING[f"obscanastillas{group}"] = f"obs_canastillas_{group}"
    COLUMN_MAPPING[f"sugcanastillas{group}"] = f"sug_canastillas_{group}"

COLUMN_MAPPING.update({
    "valvulasoperativas": "valvulas_operativas",
    "valvulasnooperativas": "valvulas_no_operativas",
    "canastillasoperativas": "canastillas_operativas",
    "canastillasnooperativas": "canastillas_no_operativas",
})

COLUMN_MAPPING.update({
    "numerosuministro": "suministro",
    "marcotapasanitaria": "marco_tapa",
    "medidasdiametro": "medidas_diametro",
    "diametrom": "medidas_diametro",
    "medidasdiametrointerno": "medidas_diametro_interno",
    "diametrointernom": "medidas_diametro_interno",
    "medidasalturautil": "medidas_altura_util",
    "alturautilm": "medidas_altura_util",
    "medidasalturatotal": "medidas_altura_total",
    "alturatotalm": "medidas_altura_total",
})

INSPECTION_ALIAS_GROUPS = {
    "caja_registro": ("cajaregistro", "cajaderegistro"),
    "marco_tapa": ("marcotapa", "marcoytapa", "marcoytapasanitaria"),
    "escalera_interior": ("escalerainterior", "escaleraint"),
    "escalera_exterior": ("escaleraexterior", "escaleraext"),
    "cuba_interior": ("cubainterior", "cubaint"),
    "cuba_exterior": ("cubaexterior", "cubaext"),
    "loza_fondo": ("lozafondo", "lozadefondo"),
    "loza_techo_interior": ("lozatechointerior", "lozatechoint"),
    "loza_techo_exterior": ("lozatechoexterior", "lozatechoext"),
    "ducto_ventilacion": ("ductoventilacion", "ductodeventilacion", "ducto"),
    "cerco_perimetrico": ("cercoperimetrico", "cerco"),
    "descarga": ("descarga", "tuberiadescarga"),
}

INSPECTION_TEXT_TARGETS = {
    "caja_registro": ("obs_caja_registro", "sug_caja_registro"),
    "marco_tapa": ("obs_marco_tapa", "sug_marco_tapa"),
    "escalera_interior": ("obs_escalera_int", "sug_escalera_int"),
    "escalera_exterior": ("obs_escalera_ext", "sug_escalera_ext"),
    "cuba_interior": ("obs_cuba_int", "sug_cuba_int"),
    "cuba_exterior": ("obs_cuba_ext", "sug_cuba_ext"),
    "loza_fondo": ("obs_loza_fondo", "sug_loza_fondo"),
    "loza_techo_interior": ("obs_loza_techo_int", "sug_loza_techo_int"),
    "loza_techo_exterior": ("obs_loza_techo_ext", "sug_loza_techo_ext"),
    "ducto_ventilacion": ("obs_ducto", "sug_ducto"),
    "cerco_perimetrico": ("obs_cerco", "sug_cerco"),
    "descarga": ("obs_descarga", "sug_descarga"),
}

for field, aliases in INSPECTION_ALIAS_GROUPS.items():
    COLUMN_MAPPING.update(dict.fromkeys(aliases, field))
    obs_target, sug_target = INSPECTION_TEXT_TARGETS[field]
    for alias in aliases:
        COLUMN_MAPPING[f"obs{alias}"] = obs_target
        COLUMN_MAPPING[f"observaciones{alias}"] = obs_target
        COLUMN_MAPPING[f"sug{alias}"] = sug_target
        COLUMN_MAPPING[f"sugerencias{alias}"] = sug_target
        COLUMN_MAPPING[f"recomendaciones{alias}"] = sug_target
        COLUMN_MAPPING[f"recomendacion{alias}"] = sug_target

VALVULA_SECTION_ALIASES = {
    "conduccion": ("conduccion", "cond"),
    "impulsion": ("impulsion", "imp"),
    "aduccion": ("aduccion",),
    "bypass": ("bypass", "pass"),
    "desague": ("desague",),
}

for section, aliases in VALVULA_SECTION_ALIASES.items():
    for alias in aliases:
        for diameter in ["2", "3", "4", "6", "8", "10", "12"]:
            COLUMN_MAPPING[f"valvulas{alias}{diameter}"] = f"valvulas_{section}_{diameter}"
            COLUMN_MAPPING[f"valv{alias}{diameter}"] = f"valvulas_{section}_{diameter}"
    COLUMN_MAPPING[f"obsvalvulas{section}"] = f"obs_valvulas_{section}"
    COLUMN_MAPPING[f"observaciones{section}"] = f"obs_valvulas_{section}"
    COLUMN_MAPPING[f"observacionesvalvulas{section}"] = f"obs_valvulas_{section}"
    COLUMN_MAPPING[f"sugvalvulas{section}"] = f"sug_valvulas_{section}"
    COLUMN_MAPPING[f"sugerencias{section}"] = f"sug_valvulas_{section}"
    COLUMN_MAPPING[f"sugerenciasvalvulas{section}"] = f"sug_valvulas_{section}"
    COLUMN_MAPPING[f"recomendaciones{section}"] = f"sug_valvulas_{section}"
    COLUMN_MAPPING[f"recomendacionesvalvulas{section}"] = f"sug_valvulas_{section}"
    COLUMN_MAPPING[f"recomendacionvalvulas{section}"] = f"sug_valvulas_{section}"

COLUMN_MAPPING["observacionespass"] = "obs_valvulas_bypass"

for section in ["aduccion", "succion", "desague"]:
    for diameter in ["2", "3", "4", "6", "8", "10", "14"]:
        COLUMN_MAPPING[f"canastillas{section}{diameter}"] = f"canastillas_{section}_{diameter}"
        COLUMN_MAPPING[f"canast{section}{diameter}"] = f"canastillas_{section}_{diameter}"
    COLUMN_MAPPING[f"obscanastillas{section}"] = f"obs_canastillas_{section}"
    COLUMN_MAPPING[f"observacionescanastilla{section}"] = f"obs_canastillas_{section}"
    COLUMN_MAPPING[f"observacionescanastillas{section}"] = f"obs_canastillas_{section}"
    COLUMN_MAPPING[f"sugcanastillas{section}"] = f"sug_canastillas_{section}"
    COLUMN_MAPPING[f"sugerenciascanastilla{section}"] = f"sug_canastillas_{section}"
    COLUMN_MAPPING[f"sugerenciascanastillas{section}"] = f"sug_canastillas_{section}"
    COLUMN_MAPPING[f"recomendacionescanastilla{section}"] = f"sug_canastillas_{section}"
    COLUMN_MAPPING[f"recomendacionescanastillas{section}"] = f"sug_canastillas_{section}"
    COLUMN_MAPPING[f"recomendacioncanastillas{section}"] = f"sug_canastillas_{section}"

COLUMN_MAPPING.update({
    "canastillasaduccion12": "canastillas_aduccion_14",
    "observacionessuccion": "obs_canastillas_succion",
    "sugerenciassuccion": "sug_canastillas_succion",
    "recomendacionessuccion": "sug_canastillas_succion",
})


def normalize_header_value(value: str) -> str:
    if not value:
        return ""
    text = str(value).strip().lower().replace("\ufeff", "")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\b(de|del|la|el)\b", "", text)
    return re.sub(r"[\s_\.:\-°/()]+", "", text)


def normalize_csv_key(value: str) -> str:
    mapped = COLUMN_MAPPING.get(normalize_header_value(value))
    if mapped:
        return mapped
    text = str(value or "").strip().lower().replace("\ufeff", "")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def import_reports_from_bytes(filename: str, content: bytes) -> list[dict[str, Any]]:
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        rows = parse_csv_file(content)
    elif lower_name.endswith(".xlsx"):
        rows = parse_xlsx_file(content)
    else:
        msg = "Formato no soportado. Use archivos .csv o .xlsx"
        raise ValueError(msg)
    if not rows:
        msg = "El archivo esta vacio o no tiene datos validos"
        raise ValueError(msg)
    reports: list[dict[str, Any]] = []
    next_report_number = 1
    for row in rows:
        explicit_number = _safe_int(row.get("informe_id"), 0)
        report_number = explicit_number if explicit_number > 0 else next_report_number
        reports.append(transform_flat_to_nested(row, report_number))
        next_report_number = max(next_report_number, report_number) + 1
    return reports


def parse_csv_file(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    sample = text[:2048]
    delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    rows: list[dict[str, Any]] = []
    for row in reader:
        normalized = {normalize_csv_key(key): value for key, value in row.items() if key is not None}
        if any(str(value or "").strip() for value in normalized.values()):
            rows.append(normalized)
    return rows


def parse_xlsx_file(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    if worksheet is None:
        return []
    rows_iter = worksheet.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return []
    keys = [normalize_csv_key(str(header or "")) for header in headers]
    rows: list[dict[str, Any]] = []
    for values in rows_iter:
        row = {keys[idx]: values[idx] for idx in range(min(len(keys), len(values))) if keys[idx]}
        if any(str(value or "").strip() for value in row.values()):
            rows.append(row)
    return rows


def transform_flat_to_nested(row: dict[str, Any], fallback_report_number: int = 1) -> dict[str, Any]:
    informe_id = _safe_int(row.get("informe_id"), 0)
    if informe_id <= 0:
        informe_id = fallback_report_number
    report = create_empty_report(informe_id)
    report["id"] = report_id_from_number(informe_id)
    report["metadata"].update({
        "informe_id": informe_id,
        "dia": _safe_int(row.get("dia"), 1),
        "mes": _resolve_mes(row.get("mes")),
        "anio": _safe_int(row.get("anio"), datetime.now().year),
        "pagina": "1 de 2",
    })
    report["header"].update({
        "cs": _safe_str(row.get("cs")),
        "contratista": _safe_str(row.get("contratista")),
        "sgio": _safe_str(row.get("sgio")),
        "codigo_infraestructura": _safe_str(row.get("codigo_infraestructura")),
        "ubicacion": _safe_str(row.get("ubicacion")),
        "suministro": _safe_str(row.get("suministro")),
        "tipo": _safe_str(row.get("tipo"), "ELEVADO").upper(),
        "volumen": _safe_int(row.get("volumen"), 0),
    })
    for item in [
        "caja_registro",
        "marco_tapa",
        "escalera_interior",
        "escalera_exterior",
        "cuba_interior",
        "cuba_exterior",
        "loza_fondo",
        "loza_techo_interior",
        "loza_techo_exterior",
        "ducto_ventilacion",
        "cerco_perimetrico",
        "descarga",
    ]:
        report["inspeccion"][item] = parse_check(row.get(item))
    _copy_text_pairs(row, report)
    _copy_valvulas(row, report)
    _copy_canastillas(row, report)
    report["medidas"].update({
        "diametro": _safe_str(row.get("medidas_diametro")),
        "diametro_interno": _safe_str(row.get("medidas_diametro_interno")),
        "altura_util": _safe_str(row.get("medidas_altura_util")),
        "altura_total": _safe_str(row.get("medidas_altura_total")),
    })
    report["observaciones"] = _safe_str(row.get("observaciones"))
    report["sugerencias"] = _safe_str(row.get("sugerencias"))
    report["last_modified"] = datetime.now().isoformat()
    return TechnicalReport.normalize(report)


def parse_check(value: Any) -> str:
    if value is None or str(value).strip() == "":
        return "unchecked"
    text = str(value).strip().upper()
    if text in {"X", "NORMAL", "BUENO", "OK", "SI", "SÍ", "V", "BIEN", "B", "N", "BUEN ESTADO"}:
        return "normal"
    if text in {"CRITICO", "CRÍTICO", "MALO", "OBSERVADO", "F", "NO", "MAL", "C", "M", "DEFICIENTE", "DAÑADO"}:
        return "critico"
    return "unchecked"


def _copy_text_pairs(row: dict[str, Any], report: dict[str, Any]) -> None:
    mapping = {
        "obs_caja_registro": "observaciones_caja_registro",
        "sug_caja_registro": "sugerencias_caja_registro",
        "obs_marco_tapa": "observaciones_marco_tapa",
        "sug_marco_tapa": "sugerencias_marco_tapa",
        "obs_escalera_int": "observaciones_escalera_int",
        "sug_escalera_int": "sugerencias_escalera_int",
        "obs_escalera_ext": "observaciones_escalera_ext",
        "sug_escalera_ext": "sugerencias_escalera_ext",
        "obs_cuba_int": "observaciones_cuba_int",
        "sug_cuba_int": "sugerencias_cuba_int",
        "obs_cuba_ext": "observaciones_cuba_ext",
        "sug_cuba_ext": "sugerencias_cuba_ext",
        "obs_loza_fondo": "observaciones_loza_fondo",
        "sug_loza_fondo": "sugerencias_loza_fondo",
        "obs_loza_techo_int": "observaciones_loza_techo_int",
        "sug_loza_techo_int": "sugerencias_loza_techo_int",
        "obs_loza_techo_ext": "observaciones_loza_techo_ext",
        "sug_loza_techo_ext": "sugerencias_loza_techo_ext",
        "obs_ducto": "observaciones_ducto",
        "sug_ducto": "sugerencias_ducto",
        "obs_cerco": "observaciones_cerco",
        "sug_cerco": "sugerencias_cerco",
        "obs_descarga": "observaciones_descarga",
        "sug_descarga": "sugerencias_descarga",
    }
    for source, target in mapping.items():
        report["inspeccion"][target] = _safe_str(row.get(source))


def _copy_valvulas(row: dict[str, Any], report: dict[str, Any]) -> None:
    section_map = {"conduccion": "diametros", "impulsion": "impulsion", "aduccion": "aduccion", "bypass": "bypass", "desague": "desague"}
    for source_section, target_section in section_map.items():
        for diameter in ["2", "3", "4", "6", "8", "10", "12"]:
            report["valvulas"][target_section][diameter] = _safe_int(row.get(f"valvulas_{source_section}_{diameter}"), 0)
        report["valvulas"][f"observaciones_{source_section}"] = _safe_str(row.get(f"obs_valvulas_{source_section}"))
        report["valvulas"][f"sugerencias_{source_section}"] = _safe_str(row.get(f"sug_valvulas_{source_section}"))
    report["valvulas"]["operativas"] = _safe_int(row.get("valvulas_operativas"), 0)
    report["valvulas"]["no_operativas"] = _safe_int(row.get("valvulas_no_operativas"), 0)


def _copy_canastillas(row: dict[str, Any], report: dict[str, Any]) -> None:
    for section in ["aduccion", "succion", "desague"]:
        for diameter in ["2", "3", "4", "6", "8", "10", "14"]:
            report["canastillas"][section][diameter] = _safe_int(row.get(f"canastillas_{section}_{diameter}"), 0)
        report["canastillas"][f"observaciones_{section}"] = _safe_str(row.get(f"obs_canastillas_{section}"))
        report["canastillas"][f"sugerencias_{section}"] = _safe_str(row.get(f"sug_canastillas_{section}"))
    for diameter in ["2", "3", "4", "6", "8", "10", "14"]:
        report["canastillas"]["diametros"][diameter] = sum(
            report["canastillas"][section][diameter] for section in ["aduccion", "succion", "desague"]
        )
    report["canastillas"]["operativas"] = _safe_int(row.get("canastillas_operativas"), 0)
    report["canastillas"]["no_operativas"] = _safe_int(row.get("canastillas_no_operativas"), 0)


def _resolve_mes(value: Any) -> str:
    if value is None or value == "":
        return "ENERO"
    try:
        number = int(float(value))
        if number in MESES:
            return MESES[number]
    except (TypeError, ValueError):
        pass
    return str(value).strip().upper()


