from io import BytesIO

from openpyxl import Workbook

from backend.core.technical_reports.importer import import_reports_from_bytes, normalize_header_value


def test_csv_semicolon_import_maps_human_headers():
    content = (
        b"Nro Informe;Centro de Servicio;Codigo Infraestructura;Tipo;Volumen;Caja Registro;Mes\n"
        b"3;SUR;RES-01;ELEVADO;150;X;5\n"
    )

    reports = import_reports_from_bytes("datos.csv", content)

    assert reports[0]["id"] == "RPT-0003"
    assert reports[0]["header"]["cs"] == "SUR"
    assert reports[0]["metadata"]["mes"] == "MAYO"
    assert reports[0]["inspeccion"]["caja_registro"] == "normal"


def test_csv_import_assigns_unique_ids_when_missing_report_number():
    content = (
        b"Centro de Servicio;Codigo Infraestructura;Tipo\n"
        b"SUR;RES-01;ELEVADO\n"
        b"NORTE;CIS-02;CISTERNA\n"
    )

    reports = import_reports_from_bytes("datos.csv", content)

    assert [report["id"] for report in reports] == ["RPT-0001", "RPT-0002"]
    assert [report["metadata"]["informe_id"] for report in reports] == [1, 2]


def test_xlsx_import_reads_rows():
    wb = Workbook()
    ws = wb.active
    ws.append(["Informe", "CS", "Codigo", "Tipo", "Volumen", "Descarga"])
    ws.append([4, "NORTE", "CIS-02", "CISTERNA", 80, "MALO"])
    buf = BytesIO()
    wb.save(buf)

    reports = import_reports_from_bytes("datos.xlsx", buf.getvalue())

    assert reports[0]["id"] == "RPT-0004"
    assert reports[0]["header"]["tipo"] == "CISTERNA"
    assert reports[0]["inspeccion"]["descarga"] == "critico"


def test_xlsx_import_maps_reference_human_header_aliases():
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Informe",
        "Numero Suministro",
        "Marco y Tapa Sanitaria",
        "Observaciones Marco y Tapa Sanitaria",
        "Sugerencias Marco y Tapa Sanitaria",
        "Valv Cond 2",
        "Canastillas Aduccion 12",
        "Observaciones Canastilla Succion",
        "Sugerencias Succion",
    ])
    ws.append([8, "NIS-99", "MALO", "CORROIDA", "CAMBIAR", 2, 1, "CON OXIDO", "MANTENIMIENTO"])
    buf = BytesIO()
    wb.save(buf)

    reports = import_reports_from_bytes("referencia.xlsx", buf.getvalue())

    report = reports[0]
    assert report["header"]["suministro"] == "NIS-99"
    assert report["inspeccion"]["marco_tapa"] == "critico"
    assert report["inspeccion"]["observaciones_marco_tapa"] == "CORROIDA"
    assert report["inspeccion"]["sugerencias_marco_tapa"] == "CAMBIAR"
    assert report["valvulas"]["diametros"]["2"] == 2
    assert report["canastillas"]["aduccion"]["14"] == 1
    assert report["canastillas"]["observaciones_succion"] == "CON OXIDO"
    assert report["canastillas"]["sugerencias_succion"] == "MANTENIMIENTO"


def test_xlsx_import_maps_recommendation_headers_to_sugerencias():
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Informe",
        "Recomendaciones Marco y Tapa Sanitaria",
        "Recomendaciones Valvulas Desague",
        "Recomendaciones Canastillas Succion",
        "Recomendaciones",
    ])
    ws.append([9, "CAMBIAR TAPA", "MANTENIMIENTO", "INSTALAR", "REVISAR EN CAMPO"])
    buf = BytesIO()
    wb.save(buf)

    reports = import_reports_from_bytes("recomendaciones.xlsx", buf.getvalue())

    report = reports[0]
    assert report["inspeccion"]["sugerencias_marco_tapa"] == "CAMBIAR TAPA"
    assert report["valvulas"]["sugerencias_desague"] == "MANTENIMIENTO"
    assert report["canastillas"]["sugerencias_succion"] == "INSTALAR"
    assert report["sugerencias"] == "REVISAR EN CAMPO"


def test_normalize_header_value_removes_accents_and_separators():
    assert normalize_header_value("Código de Infraestructura") == "codigoinfraestructura"
